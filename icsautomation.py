import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import zipfile
import os
import shutil
import tempfile
import io
import urllib.parse

# 页面配置
st.set_page_config(page_title="ICS2业务自动化整合工具", layout="wide")

# ==========================================
# 🌟 终极增强版：鼠标点击特效 (全局挂载模式)
# ==========================================
def add_click_effect():
    # 使用 components.html 或直接 markdown 注入，并增加 window 级别的持久化
    st.markdown("""
    <div id="click-effect-root"></div>
    <script>
        (function() {
            var words = ["富强", "民主", "文明", "和谐", "自由", "平等", "公正", "法治", "爱国", "敬业", "诚信", "友善"];
            var index = 0;
            
            // 绑定到最顶层窗口，确保全覆盖
            document.addEventListener('click', function(e) {
                var target = e.target;
                // 排除点击上传按钮等交互组件时可能产生的干扰
                
                var span = document.createElement("span");
                span.textContent = words[index];
                index = (index + 1) % words.length;
                
                // 随机颜色
                var color = "rgb(" + ~~(255 * Math.random()) + "," + ~~(255 * Math.random()) + "," + ~~(255 * Math.random()) + ")";
                
                // 坐标位置
                var x = e.clientX;
                var y = e.clientY;
                
                span.style.cssText = "z-index: 2147483647; position: fixed; font-weight: bold; font-family: sans-serif; pointer-events: none; user-select: none;";
                span.style.color = color;
                span.style.left = (x - 20) + "px";
                span.style.top = (y - 20) + "px";
                span.style.fontSize = "18px";
                
                document.body.appendChild(span);
                
                var top = y - 20;
                var opacity = 1;
                var scale = 1;
                
                var anim = setInterval(function() {
                    if (opacity <= 0) {
                        span.remove();
                        clearInterval(anim);
                    } else {
                        top -= 0.8;
                        opacity -= 0.015;
                        scale += 0.005;
                        span.style.top = top + "px";
                        span.style.opacity = opacity;
                        span.style.transform = "scale(" + scale + ")";
                    }
                }, 12);
            });
        })();
    </script>
    """, unsafe_allow_html=True)

# 启动特效
add_click_effect()
# ==========================================

# --- 1. 主界面标题与说明 ---
st.title("📂 ICS2业务自动化整合工具")
st.info(f"""
**· 文件一：containerinformation.xlsx**
- **来源**：此文件通常由系统生成，用于补料。
- **内容要求**：请将附件中的品名、HS CODE、件数、重量(KGS)、体积(CBM)及单号信息准确复制至本文件。其中“单号”栏位请填写客户申报ICS所用的号码；如客户未指定，则默认使用我司单号。

**· 文件二：icstemplate.xlsx**
- **来源**：此为ICS申报系统提供的标准申报表模板。
- **内容要求**：根据您所申报的主单及柜号，将相应的基础柜子信息填写完整即可。

**· 文件三：realdoc.zip**
- **来源**：此文件为客户在对单过程中，根据我司（深圳）格式要求填写的真实收发货人信息表。
- **内容要求**：收到客户回传的表格后，请先核验信息无误。随后，按申报单号将同一柜子所有客户的相关资料整理并压缩为一个ZIP文件。

:red[2026/02/14更新：解决 realdoc 无法置换问题：支持子文件夹搜索，并将空值置换为 N/A。]
""")

# --- 2. 核心处理逻辑 ---
def process_logic():
    col_up1, col_up2, col_up3 = st.columns(3)
    with col_up1:
        container_file = st.file_uploader("1. 上传 containerinformation", type=["xlsx"])
    with col_up2:
        template_file = st.file_uploader("2. 上传 icstemplate", type=["xlsx"])
    with col_up3:
        realdoc_zip = st.file_uploader("3. 上传 realdoc.zip", type=["zip"])

    if st.button("🔥 执行全流程处理", use_container_width=True):
        if not (container_file and template_file and realdoc_zip):
            st.error("请确保三个必要文件均已上传！")
            return

        with st.spinner("程序正在进行深度匹配与数据置换..."):
            try:
                with tempfile.TemporaryDirectory() as tmp_dir:
                    r_dir = os.path.join(tmp_dir, "R"); p_dir = os.path.join(tmp_dir, "P"); out_dir = os.path.join(tmp_dir, "Output")
                    os.makedirs(r_dir); os.makedirs(p_dir); os.makedirs(out_dir)

                    df = pd.read_excel(container_file)
                    df['单号'] = df['单号'].astype(str).str.strip().ffill()
                    template_bytes = template_file.read()
                    grouped = df.groupby('单号')

                    for bill_no, group in grouped:
                        if bill_no.lower() == "nan" or not bill_no: continue
                        wb = load_workbook(io.BytesIO(template_bytes))
                        ws = wb.active
                        ws['B5'] = bill_no
                        ws['B8'] = group['件数'].sum(); ws['B9'] = group['重量(KGS)'].sum()
                        f130_val = ws['F130'].value
                        for i, (_, row) in enumerate(group.iterrows()):
                            curr_row = 130 + i
                            ws[f'A{curr_row}'] = row['HS CODE']; ws[f'B{curr_row}'] = row['品名']
                            ws[f'C{curr_row}'] = row['件数']; ws[f'D{curr_row}'] = "PK-Package"
                            ws[f'E{curr_row}'] = row['重量(KGS)']; ws[f'F{curr_row}'] = f130_val
                        wb.save(os.path.join(p_dir, f"{bill_no}.xlsx"))

                    with zipfile.ZipFile(realdoc_zip, 'r') as z: z.extractall(r_dir)
                    realdoc_map = {}
                    for root, _, files in os.walk(r_dir):
                        for f in files:
                            if f.endswith('.xlsx'): realdoc_map[f.strip()] = os.path.join(root, f)

                    row_mapping = {7: 14, 8: 15, 10: 18, 11: 19}
                    columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

                    p_files = [f for f in os.listdir(p_dir) if f.endswith('.xlsx')]
                    match_count = 0
                    for filename in p_files:
                        p_path = os.path.join(p_dir, filename); r_path = realdoc_map.get(filename)
                        if r_path and os.path.exists(r_path):
                            wb_p = load_workbook(p_path); ws_p = wb_p.active
                            wb_r = load_workbook(r_path, data_only=True); ws_r = wb_r.active
                            for src_row, tgt_row in row_mapping.items():
                                for col in columns:
                                    raw_val = ws_r[f"{col}{src_row}"].value
                                    ws_p[f"{col}{tgt_row}"].value = "N/A" if (raw_val is None or str(raw_val).strip() == "") else raw_val
                            wb_p.save(os.path.join(out_dir, filename))
                            match_count += 1
                        else: shutil.copy(p_path, os.path.join(out_dir, filename))

                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as z:
                        for f in os.listdir(out_dir): z.write(os.path.join(out_dir, f), arcname=f)
                    
                    st.success(f"✅ 处理成功！已生成 {len(p_files)} 个文件。")
                    st.download_button(label="📥 下载最终结果压缩包", data=zip_buffer.getvalue(), file_name="ICS2_Results.zip", mime="application/zip", use_container_width=True)
            except Exception as e: st.error(f"⚠️ 错误: {str(e)}")

process_logic()

# --- 3. 底部：资源与支持区块 ---
st.markdown("<br><br><hr>", unsafe_allow_html=True)
st.subheader("🛠️ 资源与支持")

st.markdown("""
    <style>
    .stButton button, .stDownloadButton button, .stLinkButton a {
        color: white !important;
        font-weight: bold !important;
    }
    </style>
""", unsafe_allow_html=True)

footer_col1, footer_col2, footer_col3 = st.columns(3)

with footer_col1:
    st.markdown("#### 📖 操作指导")
    guide_path = "ICS2业务自动化整合工具使用说明.docx"
    if os.path.exists(guide_path):
        with open(guide_path, "rb") as f:
            st.download_button(label="📥 下载《使用指南.docx》", data=f.read(), file_name="ICS2业务自动化整合工具使用说明.docx", mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document", use_container_width=True)
    else: st.button("❌ 指南缺失", use_container_width=True, disabled=True)

with footer_col2:
    st.markdown("#### 📥 模板下载")
    template_files = ["containerinformation.xlsx", "icstemplate.xlsx", "realdoc.zip"]
    valid_templates = [f for f in template_files if os.path.exists(f)]
    if valid_templates:
        template_zip = io.BytesIO()
        with zipfile.ZipFile(template_zip, "w") as z:
            for f in valid_templates: z.write(f)
        st.download_button(label="📦 下载全套业务模板.zip", data=template_zip.getvalue(), file_name="ICS2_Business_Templates.zip", mime="application/zip", use_container_width=True)
    else: st.button("❌ 模板缺失", use_container_width=True, disabled=True)

with footer_col3:
    st.markdown("#### 📧 意见反馈")
    feedback_email = "yjfk@tswcbyy.com"
    st.link_button(label=f"📩 发送邮件反馈", url=f"mailto:{feedback_email}?subject=ICS2工具意见反馈", use_container_width=True)

# --- 4. 极简结尾标语 ---
st.markdown("<br><br><div style='text-align: center; color: #555e6d; font-family: sans-serif;'><p style='font-size: 18px; font-weight: 500;'>小事找GARY，大事请Google</p></div>", unsafe_allow_html=True)
